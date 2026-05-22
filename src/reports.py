import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.config import REPORTS_DIR
from src.db import get_all_expenses, get_all_income

def generate_financial_excel() -> dict:
    """
    Genera un archivo Excel estructurado y estilizado a partir de los ingresos y gastos registrados.
    
    Retorna un diccionario con el estado, la ruta del archivo generado y un mensaje descriptivo.
    """
    expenses = get_all_expenses()
    income = get_all_income()
    
    if not expenses and not income:
        return {
            "status": "empty",
            "file_path": None,
            "message": "No hay transacciones (gastos ni ingresos) registradas para generar un reporte."
        }
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"reporte_financiero_{timestamp}.xlsx"
    file_path = REPORTS_DIR / file_name
    
    # Iniciar ExcelWriter con openpyxl
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        
        # --- 1. HOJA DE RESUMEN GENERAL ---
        total_exp = sum(e['amount'] for e in expenses)
        total_inc = sum(i['amount'] for i in income)
        net_balance = total_inc - total_exp
        
        summary_data = [
            {"Concepto Financiero": "Total Ingresos", "Monto": total_inc},
            {"Concepto Financiero": "Total Gastos", "Monto": total_exp},
            {"Concepto Financiero": "Balance General", "Monto": net_balance}
        ]
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Resumen General", index=False)
        
        # --- 2. HOJA DE GASTOS DETALLADOS ---
        if expenses:
            df_exp = pd.DataFrame(expenses)
            df_exp = df_exp.rename(columns={
                "date_added": "Fecha",
                "category": "Categoría",
                "concept": "Concepto",
                "amount": "Monto"
            })
            df_exp["Fecha"] = pd.to_datetime(df_exp["Fecha"]).dt.strftime("%Y-%m-%d %H:%M")
            df_exp["Categoría"] = df_exp["Categoría"].str.capitalize()
            df_exp.to_excel(writer, sheet_name="Detalle de Gastos", index=False)
        else:
            # Crear hoja vacía si no hay gastos
            pd.DataFrame(columns=["Fecha", "Categoría", "Concepto", "Monto"]).to_excel(writer, sheet_name="Detalle de Gastos", index=False)
            
        # --- 3. HOJA DE INGRESOS DETALLADOS ---
        if income:
            df_inc = pd.DataFrame(income)
            df_inc = df_inc.rename(columns={
                "date_added": "Fecha",
                "category": "Categoría",
                "concept": "Concepto",
                "amount": "Monto"
            })
            df_inc["Fecha"] = pd.to_datetime(df_inc["Fecha"]).dt.strftime("%Y-%m-%d %H:%M")
            df_inc["Categoría"] = df_inc["Categoría"].str.capitalize()
            df_inc.to_excel(writer, sheet_name="Detalle de Ingresos", index=False)
        else:
            # Crear hoja vacía si no hay ingresos
            pd.DataFrame(columns=["Fecha", "Categoría", "Concepto", "Monto"]).to_excel(writer, sheet_name="Detalle de Ingresos", index=False)
            
        # Obtener libro de openpyxl para estilizar
        workbook = writer.book
        
        # Estilos premium compartidos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro elegante
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        green_font = Font(name="Segoe UI", size=10, bold=True, color="1E4620")
        red_font = Font(name="Segoe UI", size=10, bold=True, color="7F1D1D")
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        border_total = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        # --- ESTILIZAR HOJA: Resumen General ---
        ws_sum = workbook["Resumen General"]
        ws_sum.views.sheetView[0].showGridLines = True
        
        # Encabezados
        for col_idx in range(1, 3):
            cell = ws_sum.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin
            
        # Estilo de filas del resumen
        ws_sum.cell(row=2, column=1).font = regular_font # Total Ingresos
        ws_sum.cell(row=2, column=1).alignment = align_left
        ws_sum.cell(row=2, column=1).border = border_thin
        ws_sum.cell(row=2, column=2).font = green_font
        ws_sum.cell(row=2, column=2).alignment = align_right
        ws_sum.cell(row=2, column=2).number_format = '$#,##0.00'
        ws_sum.cell(row=2, column=2).border = border_thin
        
        ws_sum.cell(row=3, column=1).font = regular_font # Total Gastos
        ws_sum.cell(row=3, column=1).alignment = align_left
        ws_sum.cell(row=3, column=1).border = border_thin
        ws_sum.cell(row=3, column=2).font = red_font
        ws_sum.cell(row=3, column=2).alignment = align_right
        ws_sum.cell(row=3, column=2).number_format = '$#,##0.00'
        ws_sum.cell(row=3, column=2).border = border_thin
        
        # Balance General (Estilo total resaltado)
        ws_sum.cell(row=4, column=1).font = bold_font
        ws_sum.cell(row=4, column=1).alignment = align_left
        ws_sum.cell(row=4, column=1).border = border_total
        balance_cell = ws_sum.cell(row=4, column=2)
        balance_cell.font = green_font if net_balance >= 0 else red_font
        balance_cell.alignment = align_right
        balance_cell.number_format = '$#,##0.00'
        balance_cell.border = border_total
        
        ws_sum.column_dimensions['A'].width = 22
        ws_sum.column_dimensions['B'].width = 18
        
        # --- ESTILIZAR HOJAS DE DETALLE (Gastos e Ingresos) ---
        for sheet_name in ["Detalle de Gastos", "Detalle de Ingresos"]:
            ws = workbook[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            
            # Encabezados
            for col_idx in range(1, 5):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = border_thin
                
            max_row = ws.max_row
            if max_row >= 2:
                # Dar formato a las filas
                for row_idx in range(2, max_row + 1):
                    ws.cell(row=row_idx, column=1).alignment = align_center # Fecha
                    ws.cell(row=row_idx, column=2).alignment = align_left   # Categoría
                    ws.cell(row=row_idx, column=3).alignment = align_left   # Concepto
                    
                    m_cell = ws.cell(row=row_idx, column=4)
                    m_cell.alignment = align_right
                    m_cell.number_format = '$#,##0.00'
                    
                    if sheet_name == "Detalle de Ingresos":
                        m_cell.font = Font(name="Segoe UI", size=10, color="1E4620")
                    else:
                        m_cell.font = regular_font
                        
                    for col_idx in range(1, 5):
                        c = ws.cell(row=row_idx, column=col_idx)
                        if col_idx != 4 or sheet_name != "Detalle de Ingresos":
                            c.font = regular_font
                        c.border = border_thin
                        
                # Fila de totales al final
                tot_row = max_row + 1
                ws.cell(row=tot_row, column=3, value="Total").font = bold_font
                ws.cell(row=tot_row, column=3).alignment = align_right
                ws.cell(row=tot_row, column=3).border = Border(top=Side(style='thin', color='000000'))
                
                tot_val = ws.cell(row=tot_row, column=4, value=f"=SUM(D2:D{max_row})")
                tot_val.font = bold_font
                tot_val.alignment = align_right
                tot_val.number_format = '$#,##0.00'
                tot_val.border = border_total
                
            # Ajustar anchos
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value and not str(cell.value).startswith("="):
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
                
    return {
        "status": "success",
        "file_path": str(file_path),
        "message": f"Reporte financiero general generado exitosamente: {file_name}"
    }
